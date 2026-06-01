from calendar import monthrange
from datetime import date, datetime

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from core.models import (
    Category,
    Manufacturer,
    Order,
    OrderItem,
    OrderStatus,
    PickupPoint,
    Product,
    Role,
    Supplier,
    Unit,
    User,
)


# На экзамене сначала меняй этот блок: добавляй и удаляй Excel-файлы.
PICKUP_POINTS_FILE = "Пункты выдачи_import.xlsx"
PRODUCTS_FILE = "Tovar.xlsx"
USERS_FILE = "user_import.xlsx"
ORDERS_FILE = "Заказ_import.xlsx"

DATE_FORMATS = ("%d.%m.%Y", "%Y-%m-%d")


def cell(row, column):
    if column >= len(row):
        return None

    return row[column]


def text(value):
    if value is None:
        return ""

    return str(value).replace("\xa0", " ").strip()


def number(value, default=0):
    if value in (None, ""):
        return default

    return value


def excel_date(value):
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    value_text = text(value)

    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(value_text, date_format).date()
        except ValueError:
            pass

    parts = value_text.split(".")

    if len(parts) == 3 and all(part.isdigit() for part in parts):
        day = int(parts[0])
        month = int(parts[1])
        year = int(parts[2])

        if 1 <= month <= 12:
            last_day = monthrange(year, month)[1]
            return date(year, month, min(day, last_day))

    return None


def split_name(full_name):
    parts = text(full_name).split()
    last_name = parts[0] if len(parts) > 0 else ""
    first_name = parts[1] if len(parts) > 1 else ""
    patronymic = " ".join(parts[2:]) if len(parts) > 2 else ""
    return last_name, first_name, patronymic


class Command(BaseCommand):
    def handle(self, *args, **options):
        folder = settings.BASE_DIR / "import"
        if not folder.exists():
            folder = settings.BASE_DIR / "core" / "import"

        # Здесь задается порядок импорта. Справочники должны идти раньше таблиц,
        # которые на них ссылаются.
        self.import_if_exists(folder, PICKUP_POINTS_FILE, self.import_pickup_points)
        self.import_if_exists(folder, PRODUCTS_FILE, self.import_products)
        self.import_if_exists(folder, USERS_FILE, self.import_users)
        self.import_if_exists(folder, ORDERS_FILE, self.import_orders)

        self.stdout.write(self.style.SUCCESS("Импорт завершен"))
        self.stdout.write(f"Товаров: {Product.objects.count()}")
        self.stdout.write(f"Заказов: {Order.objects.count()}")
        self.stdout.write(f"Позиций заказов: {OrderItem.objects.count()}")

    def import_if_exists(self, folder, filename, import_function):
        file_path = folder / filename

        if not file_path.exists():
            self.stdout.write(f"Файл {filename} не найден, импорт пропущен")
            return

        import_function(file_path)

    def rows(self, file_path, start=2):
        sheet = load_workbook(file_path, data_only=True).active
        return sheet.iter_rows(min_row=start, values_only=True)

    def import_pickup_points(self, file_path):
        for row in self.rows(file_path, start=1):
            # Пункты выдачи_import.xlsx:
            # 0 - адрес
            address = text(cell(row, 0))

            if address:
                PickupPoint.objects.get_or_create(address=address)

    def import_products(self, file_path):
        for row in self.rows(file_path):
            # Tovar.xlsx:
            # 0 - артикул, 1 - название, 2 - единица, 3 - цена
            # 4 - поставщик, 5 - производитель, 6 - категория
            # 7 - скидка, 8 - количество, 9 - описание, 10 - фото
            article = text(cell(row, 0))
            name = text(cell(row, 1))
            unit_name = text(cell(row, 2))
            price = number(cell(row, 3))
            supplier_name = text(cell(row, 4))
            manufacturer_name = text(cell(row, 5))
            category_name = text(cell(row, 6))
            discount = number(cell(row, 7))
            quantity = number(cell(row, 8))
            description = text(cell(row, 9))
            photo = text(cell(row, 10))

            if not article:
                continue

            supplier, _ = Supplier.objects.get_or_create(name=supplier_name)
            manufacturer, _ = Manufacturer.objects.get_or_create(name=manufacturer_name)
            category, _ = Category.objects.get_or_create(name=category_name)
            unit, _ = Unit.objects.get_or_create(name=unit_name or "шт.")

            Product.objects.update_or_create(
                article=article,
                defaults={
                    "name": name,
                    "unit": unit,
                    "price": price,
                    "supplier": supplier,
                    "manufacturer": manufacturer,
                    "category": category,
                    "discount": discount,
                    "quantity": quantity,
                    "description": description,
                    "photo": f"products/{photo}" if photo else "",
                },
            )

    def import_users(self, file_path):
        for row in self.rows(file_path):
            # user_import.xlsx:
            # 0 - роль, 1 - ФИО, 2 - логин, 3 - пароль
            role_from_excel = text(cell(row, 0))
            full_name = cell(row, 1)
            login = text(cell(row, 2))
            password = text(cell(row, 3))
            last_name, first_name, patronymic = split_name(full_name)

            if not login:
                continue

            role_name = role_from_excel
            if role_from_excel == "Администратор":
                role_name = "admin"
            elif role_from_excel == "Менеджер":
                role_name = "manager"
            elif role_from_excel == "Авторизированный клиент":
                role_name = "client"

            role, _ = Role.objects.get_or_create(name=role_name)

            user, _ = User.objects.update_or_create(
                username=login,
                defaults={
                    "email": login,
                    "last_name": last_name,
                    "first_name": first_name,
                    "patronymic": patronymic,
                    "role": role,
                },
            )
            user.set_password(password)
            user.save()

    def import_orders(self, file_path):
        for row in self.rows(file_path):
            # Заказ_import.xlsx:
            # 0 - ID, 1 - товары, 2 - дата заказа, 3 - дата выдачи
            # 4 - пункт выдачи, 5 - пользователь, 6 - код, 7 - статус
            order_id = cell(row, 0)
            items_text = text(cell(row, 1))
            order_date_value = cell(row, 2)
            delivery_date_value = cell(row, 3)
            pickup_point_id = cell(row, 4)
            user_full_name = cell(row, 5)
            pickup_code = number(cell(row, 6))
            status_name = text(cell(row, 7))
            order_date = excel_date(order_date_value)
            delivery_date = excel_date(delivery_date_value)

            if not order_id:
                continue

            if not order_date or not delivery_date:
                self.stdout.write(f"Заказ {order_id} пропущен: неправильная дата")
                continue

            pickup_point = PickupPoint.objects.filter(id=pickup_point_id).first()
            if pickup_point is None:
                pickup_point = PickupPoint.objects.first()

            status, _ = OrderStatus.objects.get_or_create(name=text(status_name))

            order, _ = Order.objects.update_or_create(
                id=order_id,
                defaults={
                    "order_date": order_date,
                    "delivery_date": delivery_date,
                    "pickup_point": pickup_point,
                    "user": self.find_user(user_full_name),
                    "pickup_code": pickup_code,
                    "status": status,
                },
            )

            OrderItem.objects.filter(order=order).delete()
            self.create_order_items(order, items_text)

    def create_order_items(self, order, items_text):
        parts = [part.strip() for part in items_text.split(",") if part.strip()]

        for index in range(0, len(parts), 2):
            if index + 1 >= len(parts):
                continue

            product = Product.objects.filter(article=parts[index]).first()

            if product:
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    count=int(parts[index + 1]),
                )

    def find_user(self, full_name):
        last_name, first_name, patronymic = split_name(full_name)
        return User.objects.filter(
            last_name=last_name,
            first_name=first_name,
            patronymic=patronymic,
        ).first()
